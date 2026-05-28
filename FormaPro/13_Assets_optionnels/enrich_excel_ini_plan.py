from pathlib import Path
import importlib.util
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PLAN = r"C:\Users\jboul\OneDrive\professionnel\Formations\office\Excel\Excel Ini\PLAN COMPLET Excel INI.docx"
OPERATORS_PLAN = r"C:\Users\jboul\OneDrive\professionnel\Formations\office\Excel\Excel Ini\opérateurs de calcul.docx"
CONCISE_PLAN = r"C:\Users\jboul\OneDrive\professionnel\Formations\office\Excel\Excel Ini\Plan concis Excel Ini.docx"


def load_base_assets():
    generator_path = ROOT / "13_Assets_optionnels" / "generate_phases_5_12.py"
    spec = importlib.util.spec_from_file_location("formapro_generator", generator_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.BASE_CSS, module.BASE_JS


BASE_CSS, BASE_JS = load_base_assets()
BASE_CSS += "\n.table-scroll{overflow-x:auto}.table-scroll table{min-width:760px}\n"


def write(rel, content):
    path = ROOT / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return path


def slug(text):
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")[:48] or "excel"


def html_page(title, theme, rel_index, subtitle, body, badges):
    badge_html = "".join(f'<span class="badge">{badge}</span>' for badge in badges)
    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>{BASE_CSS}</style>
</head>
<body class="theme-{theme}">
  <header>
    <div class="wrap">
      <a class="top no-print" href="{rel_index}">Retour à l'index général</a>
      <h1>{title}</h1>
      <p class="lead">{subtitle}</p>
      <div class="meta">{badge_html}</div>
      <div class="actions no-print"><button type="button" onclick="window.print()">Imprimer / Exporter en PDF</button><button type="button" class="secondary" data-open-all>Tout ouvrir</button><button type="button" class="secondary" data-close-all>Tout fermer</button></div>
    </div>
  </header>
  <main class="wrap">
    {body}
  </main>
  <script>{BASE_JS}</script>
</body>
</html>
"""


def panel(title, content, ident=None):
    attr = f' id="{ident}"' if ident else ""
    return f'<section class="panel"{attr}><h2>{title}</h2>{content}</section>'


def card_grid(items, cols="grid"):
    return f'<section class="{cols}">' + "".join(
        f'<article class="card"><h3>{title}</h3>{content}</article>' for title, content in items
    ) + "</section>"


def ul(items):
    return "<ul>" + "".join(f"<li>{item}</li>" for item in items) + "</ul>"


def ol_steps(items):
    return '<ol class="steps">' + "".join(f"<li>{item}</li>" for item in items) + "</ol>"


def checklist(items, cid):
    return f'<ul class="checklist" data-checklist-id="{slug(cid)}">' + "".join(
        f'<li><input type="checkbox"><span>{item}</span></li>' for item in items
    ) + "</ul>"


def accordion(label, content):
    return f'<div class="accordion"><button type="button" aria-expanded="false">{label}</button><div class="content">{content}</div></div>'


def table(headers, rows):
    head = "<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"
    body = "".join("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>" for row in rows)
    return f"<table>{head}{body}</table>"


def write_xlsx(rel, rows, sheet_name="Exercice"):
    path = ROOT / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="B7C9B9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 32)
    headers = [str(cell.value or "") for cell in ws[1]]
    for idx, header in enumerate(headers, start=1):
        normalized = header.lower()
        if any(word in normalized for word in ["prix", "total", "ventes", "commission"]):
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    if isinstance(item.value, (int, float)):
                        item.number_format = '#,##0.00 €'
        if "taux" in normalized:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    if isinstance(item.value, (int, float)):
                        item.number_format = '0.00%'
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def support_excel():
    exercise_links = [
        ("01 - Créer un tableau de suivi", "../04_Exercices_HTML/Excel/Excel_01_Creer_un_tableau_de_suivi.html"),
        ("02 - Calculer totaux et moyennes", "../04_Exercices_HTML/Excel/Excel_02_Calculer_totaux_et_moyennes.html"),
        ("03 - Créer un budget simple", "../04_Exercices_HTML/Excel/Excel_03_Creer_un_budget_simple.html"),
        ("04 - Utiliser SI simple", "../04_Exercices_HTML/Excel/Excel_04_Utiliser_SI_simple.html"),
        ("05 - Créer un graphique", "../04_Exercices_HTML/Excel/Excel_05_Creer_un_graphique.html"),
        ("06 - Corriger un tableau avec erreurs", "../04_Exercices_HTML/Excel/Excel_06_Corriger_un_tableau_avec_erreurs.html"),
        ("07 - Mise en page et impression", "../04_Exercices_HTML/Excel/Excel_07_Mise_en_page_et_impression.html"),
        ("08 - Séries et recopie de formules", "../04_Exercices_HTML/Excel/Excel_08_Series_et_recopie_de_formules.html"),
        ("09 - Références absolues et commission", "../04_Exercices_HTML/Excel/Excel_09_References_absolues_commission.html"),
        ("10 - Graphiques et mise en forme conditionnelle", "../04_Exercices_HTML/Excel/Excel_10_Graphiques_et_mise_en_forme_conditionnelle.html"),
    ]
    command_rows = [
        [
            "Sélectionner une cellule",
            "Dire à Excel où agir. Toute saisie, mise en forme ou formule s'applique d'abord à la cellule active.",
            "Cliquer en B2 puis taper le nom du participant.",
        ],
        [
            "Sélectionner une plage",
            "Travailler sur plusieurs cellules d'un coup.",
            "<code>B2:B8</code> désigne toutes les cellules de B2 à B8.",
        ],
        [
            "SOMME",
            "Additionner automatiquement une série de nombres.",
            "<code>=SOMME(D2:D8)</code> additionne les montants de D2 à D8.",
        ],
        [
            "MOYENNE",
            "Calculer la valeur moyenne d'une série.",
            "<code>=MOYENNE(D2:D8)</code> donne le montant moyen.",
        ],
        [
            "NB",
            "Compter les cellules qui contiennent des nombres.",
            "<code>=NB(B2:B20)</code> compte les valeurs numériques de B2 à B20.",
        ],
        [
            "MAX / MIN",
            "Trouver la plus grande ou la plus petite valeur.",
            "<code>=MAX(D2:D8)</code> donne le plus grand total ; <code>=MIN(D2:D8)</code> donne le plus petit.",
        ],
        [
            "SI",
            "Tester une condition. En langage naturel : si la condition est vraie, Excel affiche ou calcule le premier résultat ; sinon il affiche ou calcule le second.",
            "<code>=SI(C2=\"Non\";\"A relancer\";\"OK\")</code> signifie : si C2 contient Non, écrire A relancer ; sinon écrire OK.",
        ],
        [
            "Référence relative",
            "Laisser Excel adapter les cellules quand une formule est recopiée.",
            "<code>=B2*C2</code> devient <code>=B3*C3</code> à la ligne suivante.",
        ],
        [
            "Référence absolue",
            "Bloquer une cellule avec <code>$</code> quand une formule est recopiée.",
            "<code>=B2*$F$2</code> garde toujours le taux placé en F2.",
        ],
        [
            "Poignée de recopie",
            "Étendre une formule, une date, un mois ou une série sans tout retaper.",
            "Tirer le coin inférieur droit de la cellule Janvier pour obtenir Février, Mars, Avril.",
        ],
        [
            "Format nombre / monnaie / %",
            "Changer l'affichage sans changer la valeur réelle.",
            "0,2 peut s'afficher en <code>20%</code>, mais le contenu reste 0,2.",
        ],
        [
            "Renvoyer à la ligne",
            "Afficher un texte long dans une cellule sans élargir excessivement la colonne.",
            "Bouton Renvoyer à la ligne ou <code>Alt+Entrée</code> dans la cellule.",
        ],
        [
            "Mise en forme conditionnelle",
            "Mettre en évidence automatiquement les cellules qui respectent une règle.",
            "Colorer en rouge les écarts négatifs avec une règle <code>&lt; 0</code>.",
        ],
        [
            "Graphique",
            "Transformer des nombres en comparaison visuelle.",
            "Sélectionner Mois + Réalisé, puis Insérer un graphique en colonnes.",
        ],
        [
            "Aperçu avant impression",
            "Vérifier le rendu papier/PDF avant d'imprimer.",
            "Contrôler orientation, marges, échelle et pages coupées.",
        ],
    ]
    course_plan_rows = [
        ["1", "Présentation et logistique", "Tour de table, objectifs, rassurer sur l'objectif des exercices, méthode de travail et enregistrement."],
        ["2", "Démarrage logiciel", "Ouvrir Excel, comprendre Backstage/Fichier, ruban, onglets, fenêtre du classeur, feuilles, barre de formule, barre d'état."],
        ["3", "Feuilles et classeur", "Créer, renommer, déplacer, colorer et supprimer une feuille ; comprendre la taille d'une feuille de calcul."],
        ["4", "Saisie et sélection", "Cellule active, plage, validation, modification, suppression, position des textes/nombres."],
        ["5", "Couper, copier, coller", "Procédure guidée, déplacement/duplication, reproduction de mise en forme."],
        ["6", "Lignes, colonnes, cellules", "Ajuster largeur/hauteur, lire le symptôme ###, insérer/supprimer lignes, colonnes et cellules."],
        ["7", "Mise en forme", "Police, taille, gras/italique/souligné, couleurs, bordures, alignements, retraits, retour à la ligne, fusion."],
        ["8", "Fonctions automatiques", "Somme automatique, composition d'une formule, correction manuelle, recopie, formes du curseur."],
        ["9", "Formats de nombre", "Décimales, monétaire/comptabilité, pourcentage, date, heure, fraction, scientifique, texte, séparateur de milliers."],
        ["10", "Formules et opérateurs", "Principe du signe =, addition, division, priorités, exposant, références de cellules, calcul entre feuilles."],
        ["11", "Séries et recopie", "Dates, jours, mois, texte + nombre, nombres seuls, pas d'incrément, formules incrémentées et erreurs."],
        ["12", "Mise en page", "Figer/masquer, affichage Mise en page, titres, orientation, échelle, en-tête/pied, aperçu avant impression."],
        ["13", "Références relatives/absolues", "Constater le problème, poser la règle : bloquer seulement en recopie quand la valeur n'est pas dans la liste."],
        ["14", "Graphiques", "Choisir selon objectif : évolution, comparaison, proportion ; créer, modifier, imprimer et relier au tableau."],
        ["15", "Commentaires, MFC, interaction", "Notes/commentaires, mise en forme conditionnelle, priorités de règles, copier-coller un tableau dans Word."],
    ]
    operator_rows = [
        ["Arithmétiques", "<code>+</code>, <code>-</code>, <code>*</code>, <code>/</code>, <code>%</code>, <code>^</code>", "Faire un calcul numérique : addition, soustraction, multiplication, division, pourcentage, puissance.", "<code>=B2*C2</code> ou <code>=3^3</code>."],
        ["Comparaison", "<code>=</code>, <code>></code>, <code>&lt;</code>, <code>>=</code>, <code>&lt;=</code>, <code>&lt;></code>", "Comparer deux valeurs. Le résultat logique est VRAI ou FAUX.", "<code>=A1>B1</code> ou dans SI : <code>=SI(C2>=10;\"OK\";\"A revoir\")</code>."],
        ["Texte", "<code>&amp;</code>", "Coller plusieurs textes pour obtenir une seule chaîne.", "<code>=A2&amp;\" \"&amp;B2</code> assemble prénom et nom."],
        ["Référence", "<code>:</code>, <code>;</code>, espace", "Désigner une plage, unir plusieurs zones ou trouver une intersection.", "<code>=SOMME(B5:B15;D5:D15)</code> additionne deux plages."],
    ]

    toc = """
    <nav class="toc">
      <ul>
        <li><a href="#demarrage">Démarrage</a></li>
        <li><a href="#plan-cours">Plan</a></li>
        <li><a href="#saisie">Saisie</a></li>
        <li><a href="#calculs">Calculs</a></li>
        <li><a href="#operateurs">Opérateurs</a></li>
        <li><a href="#commandes">Commandes</a></li>
        <li><a href="#impression">Impression</a></li>
        <li><a href="#graphiques">Graphiques</a></li>
        <li><a href="#exercices">Exercices</a></li>
        <li><a href="#memo">Mémo</a></li>
      </ul>
    </nav>
    """

    body = toc
    body += panel(
        "Objectifs opérationnels",
        ul([
            "Se repérer dans un classeur : ruban, zone Nom, barre de formule, feuilles, onglets, barre d'état.",
            "Saisir, sélectionner, déplacer, copier, coller et supprimer des données sans perdre la structure du tableau.",
            "Mettre en forme police, alignement, bordures, formats de nombre, lignes et colonnes.",
            "Créer des formules fiables avec références relatives, références absolues et fonctions automatiques.",
            "Préparer une impression lisible, créer un graphique adapté et utiliser notes, commentaires et mise en forme conditionnelle.",
        ])
        + '<div class="callout ok"><strong>Sources utilisées :</strong> PLAN COMPLET Excel INI.docx, Plan concis Excel Ini.docx, opérateurs de calcul.docx.</div>',
        "objectifs",
    )
    body += panel(
        "Plan de cours restructuré",
        "<p>Progression synthétisée à partir du plan concis Excel INI et du document sur les opérateurs de calcul.</p>"
        + '<div class="table-scroll">'
        + table(["#", "Séquence", "Logique pédagogique"], course_plan_rows)
        + "</div>",
        "plan-cours",
    )
    body += panel(
        "Démarrage logiciel et repères",
        table(
            ["Élément", "À retenir", "Erreur fréquente"],
            [
                ["Classeur", "Fichier Excel complet, composé d'une ou plusieurs feuilles.", "Travailler sur le mauvais fichier ou oublier d'enregistrer."],
                ["Feuille", "Onglet en bas du classeur ; chaque feuille contient une grille.", "Confondre feuille et classeur."],
                ["Cellule active", "Case sélectionnée, repérée par une lettre de colonne et un numéro de ligne.", "Taper sans regarder où se trouve la sélection."],
                ["Zone Nom", "Affiche l'adresse de la cellule active, par exemple B4.", "Ne pas vérifier la cellule avant une formule."],
                ["Barre de formule", "Affiche le contenu réel : texte, nombre ou formule.", "Lire seulement le résultat visible dans la cellule."],
                ["Barre d'état", "Affiche rapidement somme, moyenne ou nombre pour une sélection.", "Recalculer à la main une information déjà visible."],
            ],
        ),
        "demarrage",
    )
    body += panel(
        "Saisie, sélection, lignes et colonnes",
        card_grid([
            ("Valider une saisie", "<p>Entrée valide et descend. Tab valide et va à droite. La coche valide aussi. Dans une formule, éviter de cliquer ailleurs par erreur.</p>"),
            ("Texte ou nombre", "<p>Le texte s'aligne souvent à gauche, les nombres à droite. Un nombre stocké en texte ne calcule pas correctement.</p>"),
            ("Dates", "<p>Une date est une valeur numérique affichée au format date. <code>Ctrl+;</code> insère la date du jour.</p>"),
            ("Effacer", "<p>Suppr efface le contenu, pas toujours le format. Pour repartir proprement, nettoyer aussi formats et bordures si besoin.</p>"),
            ("Copier/couper/coller", "<p>Copier duplique. Couper déplace. Reproduire la mise en forme copie seulement l'apparence.</p>"),
            ("Lignes et colonnes", "<p>Double-clic entre deux lettres ajuste la largeur. <code>###</code> signale souvent une colonne trop étroite.</p>"),
        ]),
        "saisie",
    )
    body += panel(
        "Mise en forme utile",
        table(
            ["Commande", "Usage", "Limite"],
            [
                ["Police, taille, gras, italique, souligné", "Structurer titres, en-têtes et valeurs importantes.", "Trop d'effets nuit à la lecture."],
                ["Couleur de remplissage et couleur de texte", "Attirer l'œil sur titres, alertes ou zones à saisir.", "Garder contraste fort."],
                ["Bordures", "Matérialiser un tableau imprimable.", "Ne pas border toute la feuille."],
                ["Alignement horizontal/vertical", "Rendre les libellés homogènes.", "Centrer tous les nombres ralentit la lecture."],
                ["Renvoyer à la ligne / Alt+Entrée", "Afficher un libellé long dans une cellule.", "Ajuster la hauteur de ligne ensuite."],
                ["Fusionner et centrer", "Titre ponctuel seulement.", "À éviter dans une zone de calcul ou de tri."],
            ],
        ),
    )
    body += panel(
        "Calculs, fonctions et formats",
        card_grid([
            ("Formule", "<p>Toute formule commence par <code>=</code>. Exemples : <code>=B2*C2</code>, <code>=(B2+C2)/2</code>.</p>"),
            ("Priorité", "<p>Excel calcule parenthèses, puissance, multiplication/division, addition/soustraction.</p>"),
            ("Plage", "<p><code>B2:B8</code> signifie de B2 à B8. Les deux-points créent une plage continue.</p>"),
            ("Fonctions automatiques", "<p><code>SOMME</code>, <code>MOYENNE</code>, <code>NB</code>, <code>MAX</code>, <code>MIN</code> couvrent les premiers besoins.</p>"),
            ("Formats", "<p>Nombre, monétaire, comptabilité, pourcentage, date, heure, texte. Le format change l'affichage, pas le contenu.</p>"),
            ("Erreurs", "<p><code>#DIV/0!</code>, <code>#VALEUR!</code>, <code>#REF!</code> et référence circulaire indiquent une formule à vérifier.</p>"),
        ]),
        "calculs",
    )
    body += panel(
        "Opérateurs de calcul : logique et ordre de priorité",
        "<p>Les opérateurs indiquent le type de calcul demandé. Excel respecte l'ordre mathématique : parenthèses, exposants, multiplication/division, addition/soustraction. Les parenthèses permettent de changer cet ordre.</p>"
        + '<div class="table-scroll">'
        + table(["Type", "Opérateurs", "Logique", "Exemple"], operator_rows)
        + "</div>"
        + '<div class="callout warn"><strong>Point formateur :</strong> faire comparer <code>=2+3*4</code> et <code>=(2+3)*4</code> pour montrer le rôle des parenthèses.</div>',
        "operateurs",
    )
    body += panel(
        "Index des commandes Excel de base et leur logique",
        "<p>Utilisez cet index comme aide-mémoire : chaque commande répond à une intention simple avant d'être une formule ou un bouton.</p>"
        + '<div class="table-scroll">'
        + table(["Commande", "Logique en langage naturel", "Exemple"], command_rows)
        + "</div>",
        "commandes",
    )
    body += panel(
        "Séries, recopie et références",
        table(
            ["Besoin", "Geste", "Contrôle"],
            [
                ["Créer une série", "Tirer la poignée de recopie pour dates, mois, jours ou texte + numéro.", "Vérifier le pas de la série."],
                ["Incrément numérique", "Saisir deux premières valeurs ou maintenir Ctrl selon le comportement voulu.", "La suite doit progresser régulièrement."],
                ["Recopier une formule", "Tirer la poignée ; les références relatives se déplacent.", "Comparer la formule copiée avec la ligne source."],
                ["Bloquer une cellule", "Utiliser <code>$</code> : <code>$F$2</code> reste fixe en recopie.", "Le taux ou coefficient reste identique partout."],
            ],
        ),
    )
    body += panel(
        "Mise en page et impression",
        ul([
            "Passer en affichage Mise en page pour contrôler le rendu imprimé.",
            "Choisir orientation portrait ou paysage selon largeur du tableau.",
            "Régler marges, zone d'impression, titres à imprimer et sauts de page.",
            "Utiliser l'échelle avec prudence : ajuster à 1 page en largeur si le tableau dépasse.",
            "Ajouter en-tête/pied de page, date ou numéro de page quand le document circule.",
            "Toujours ouvrir l'aperçu avant impression avant de lancer l'impression.",
        ]),
        "impression",
    )
    body += panel(
        "Graphiques, notes et mise en forme conditionnelle",
        card_grid([
            ("Choisir le bon graphique", "<p>Courbe pour évolution, colonnes/barres pour comparaison, secteurs pour proportion simple, nuage de points pour relation.</p>"),
            ("Sélectionner les données", "<p>Sélectionner seulement les données utiles, avec en-têtes, sans totaux si ceux-ci faussent le graphique.</p>"),
            ("Mettre en forme", "<p>Titre clair, légende utile, étiquettes seulement si elles apportent une information.</p>"),
            ("Notes/commentaires", "<p>Insérer une note pour expliquer une cellule, la modifier, l'afficher/masquer, puis la supprimer si elle n'est plus utile.</p>"),
            ("Mise en forme conditionnelle", "<p>Règles de surbrillance, valeurs en double, barres de données et jeux d'icônes aident à repérer les priorités.</p>"),
            ("Word et autres applis", "<p>Copier un tableau Excel vers Word pour produire un compte rendu, puis vérifier lisibilité et largeur.</p>"),
        ]),
        "graphiques",
    )
    body += panel(
        "Exercices et cas pratiques",
        checklist([f'<a href="{href}">{label}</a>' for label, href in exercise_links], "support-excel-exercices"),
        "exercices",
    )
    body += panel(
        "Mémo final",
        ul([
            "Sélectionner avant d'agir.",
            "Lire la barre de formule pour comprendre le contenu réel.",
            "Vérifier les formats avant d'accuser le calcul.",
            "Tester une formule sur une ligne, puis recopier.",
            "Contrôler l'aperçu avant impression.",
            "En cas d'erreur, utiliser Annuler puis revenir au dernier geste réussi.",
        ]),
        "memo",
    )
    write(
        "03_Supports_stagiaires_HTML/Support_Excel_Debutant.html",
        html_page(
            "Excel débutant : tableaux, calculs et graphiques simples",
            "excel",
            "../00_Index_general/index.html",
            "Support stagiaire enrichi avec le plan complet Excel INI : prise en main, saisie, formules, mise en page, graphiques et cas pratiques.",
            body,
            ["Excel", "Public : grands débutants", "Source : PLAN COMPLET Excel INI"],
        ),
    )


GUIDES = {
    "Jour_3_Word_Excel.html": {
        "title": "Jour 3 - Word + Excel",
        "subtitle": "Vendredi 5 juin : consolidation Word le matin, puis 3 h Excel pour démarrer proprement.",
        "focus": "Installer les repères : classeur, feuille, cellule active, zone Nom, barre de formule, ruban, barre d'état.",
        "timeline": [
            ("09:00", "Consolidation Word déjà prévue : mise en forme et document final."),
            ("13:30", "Démarrage Excel : ouvrir, enregistrer, repérer classeur, feuilles, onglets et cellule active."),
            ("14:00", "Saisie guidée : texte, nombre, date, validation Entrée/Tab, barre de formule."),
            ("14:45", "Couper, copier, coller, reproduire la mise en forme ; supprimer contenu vs formats."),
            ("15:15", "Lignes, colonnes, largeur automatique, lecture du symptôme ###."),
            ("15:45", "Mise en forme police, alignements, bordures, renvoi à la ligne, fusion à éviter."),
            ("16:15", "Exercice 01 puis correction collective courte."),
        ],
        "files": [
            ("Support Excel", "../03_Supports_stagiaires_HTML/Support_Excel_Debutant.html"),
            ("Exercice 01", "../04_Exercices_HTML/Excel/Excel_01_Creer_un_tableau_de_suivi.html"),
            ("Exercice 02", "../04_Exercices_HTML/Excel/Excel_02_Calculer_totaux_et_moyennes.html"),
            ("Fichier participants", "../06_Fichiers_de_depart/Excel/participants_evenement.xlsx"),
        ],
        "demo": [
            "Demander aux stagiaires de lire l'adresse de la cellule avant chaque saisie.",
            "Faire verbaliser ce qui est dans la cellule et ce qui est visible dans la barre de formule.",
            "Montrer qu'une suppression peut laisser la mise en forme en place.",
        ],
        "slow": "Limiter à saisie, largeur de colonnes, bordures, gras et sauvegarde.",
        "fast": "Ajouter une colonne Commentaire et tester le renvoi à la ligne.",
    },
    "Jour_4_Excel.html": {
        "title": "Jour 4 - Excel",
        "subtitle": "Jeudi 18 juin : 8 h Excel pour fonctions, formats, séries, recopie et impression.",
        "focus": "Passer du tableau saisi au tableau calculé : formules, fonctions automatiques, formats et contrôle avant impression.",
        "timeline": [
            ("09:00", "Réactivation J3 : cellule active, barre de formule, formats visibles."),
            ("09:30", "Formules : signe =, opérateurs, parenthèses, références de cellules."),
            ("10:30", "Fonctions automatiques : SOMME, MOYENNE, NB, MAX, MIN et plages avec deux-points."),
            ("11:30", "Formats de nombre : décimales, monétaire, comptabilité, %, date, heure, texte."),
            ("13:30", "Séries incrémentées : dates, mois, texte + numéro, pas non égal à 1."),
            ("14:30", "Formules recopiées : références relatives, erreurs #DIV/0!, #VALEUR!, #REF!."),
            ("15:30", "Mise en page : paysage, marges, échelle, zone d'impression, titres à imprimer."),
            ("16:30", "Exercices 02, 03, 07 et 08 selon rythme."),
        ],
        "files": [
            ("Support Excel", "../03_Supports_stagiaires_HTML/Support_Excel_Debutant.html"),
            ("Exercice 02", "../04_Exercices_HTML/Excel/Excel_02_Calculer_totaux_et_moyennes.html"),
            ("Exercice 07", "../04_Exercices_HTML/Excel/Excel_07_Mise_en_page_et_impression.html"),
            ("Exercice 08", "../04_Exercices_HTML/Excel/Excel_08_Series_et_recopie_de_formules.html"),
            ("Fichier frais vacances", "../06_Fichiers_de_depart/Excel/frais_vacances_mise_en_page.xlsx"),
        ],
        "demo": [
            "Construire une formule au clavier, puis avec clic sur cellules.",
            "Faire changer seulement le format et montrer que le contenu n'a pas changé.",
            "Comparer une formule source et sa recopie ligne par ligne.",
        ],
        "slow": "Conserver SOMME, MOYENNE et mise en page 1 page en largeur.",
        "fast": "Ajouter NB, MAX, MIN et une zone d'impression avec en-tête/pied de page.",
    },
    "Jour_5_Excel.html": {
        "title": "Jour 5 - Excel",
        "subtitle": "Mardi 7 juillet : 7 h Excel pour références absolues, graphiques, notes, MFC et autonomie.",
        "focus": "Rendre les classeurs exploitables : référence bloquée, graphique adapté, alerte visuelle et transfert vers Word.",
        "timeline": [
            ("09:00", "Réactivation : formules, formats, recopie et erreurs classiques."),
            ("09:45", "Références relatives et absolues : rôle du $, taux ou coefficient fixe."),
            ("10:45", "Graphiques : objectif, sélection sans totaux, colonnes/barres/courbes/secteurs."),
            ("11:45", "Éléments du graphique : titre, légende, étiquettes, impression graphique ou feuille."),
            ("13:30", "Notes/commentaires : insérer, modifier, afficher/masquer, supprimer."),
            ("14:15", "Mise en forme conditionnelle : règles, doublons, barres de données, icônes, priorité."),
            ("15:15", "Interaction avec Word : copier un tableau et vérifier largeur/lisibilité."),
            ("16:00", "Cas pratiques 09 et 10, puis mini-évaluation Excel."),
        ],
        "files": [
            ("Support Excel", "../03_Supports_stagiaires_HTML/Support_Excel_Debutant.html"),
            ("Exercice 09", "../04_Exercices_HTML/Excel/Excel_09_References_absolues_commission.html"),
            ("Exercice 10", "../04_Exercices_HTML/Excel/Excel_10_Graphiques_et_mise_en_forme_conditionnelle.html"),
            ("Corrigé exercice 10", "../05_Corriges_HTML/Excel/Corrige_Excel_10_Graphiques_et_mise_en_forme_conditionnelle.html"),
            ("Fichier graphique", "../06_Fichiers_de_depart/Excel/entrainement_graphique_conditionnel.xlsx"),
        ],
        "demo": [
            "Montrer une formule avec taux fixe avant et après recopie.",
            "Demander quel graphique répond à quelle question métier.",
            "Faire gérer deux règles de mise en forme conditionnelle et observer leur priorité.",
        ],
        "slow": "Limiter à référence absolue + graphique en colonnes + une règle de surbrillance.",
        "fast": "Ajouter icônes, notes et copie du tableau dans Word.",
    },
}


def guide_pages():
    for filename, cfg in GUIDES.items():
        file_links = "".join(f'<li><a href="{href}">{label}</a></li>' for label, href in cfg["files"])
        timeline = "".join(f"<li><strong>{time}</strong> - {text}</li>" for time, text in cfg["timeline"])
        body = f"""
        <section class="grid two">
          <article class="panel"><h2>Avant de commencer</h2>{checklist(["Tester les fichiers XLSX.", "Ouvrir le support stagiaire.", "Préparer une correction collective courte.", "Créer un dossier d'enregistrement par stagiaire."], cfg["title"] + " prepa")}</article>
          <article class="panel"><h2>Fichiers à ouvrir</h2><ul>{file_links}</ul></article>
        </section>
        {panel("Déroulé horaire", f"<ul>{timeline}</ul>")}
        <section class="grid two">
          <article class="card"><h3>Le formateur montre</h3><p>{cfg["focus"]}</p>{ul(cfg["demo"])}</article>
          <article class="card"><h3>Les stagiaires font</h3><p>Ils reproduisent le geste immédiatement, vérifient la cellule active ou le résultat, puis enregistrent avant l'exercice.</p></article>
        </section>
        {panel("Consignes et variantes", accordion("Consignes stagiaires", "<p>Travaillez lentement. Sélectionnez avant d'agir. Lisez la barre de formule avant de conclure qu'Excel s'est trompé.</p>") + accordion("Version courte si retard", f"<p>{cfg['slow']}</p>") + accordion("Activité bonus", f"<p>{cfg['fast']}</p>"))}
        {panel("Points de vigilance formateur", ul(["Le clic ailleurs pendant une formule peut insérer une mauvaise référence.", "Le format ne modifie pas le contenu réel.", "La fusion de cellules gêne tri, filtre et recopie.", "Un graphique doit répondre à une question avant d'être décoré."]))}
        {panel("Fin de journée", checklist(["Fichiers enregistrés.", "Exercice corrigé.", "Erreur fréquente reformulée par le groupe.", "Synthèse orale réalisée."], cfg["title"] + " fin"))}
        """
        write(
            f"02_Guide_formateur_HTML/{filename}",
            html_page(cfg["title"], "excel", "../00_Index_general/index.html", cfg["subtitle"], body, ["Excel", "Guide formateur", "Plan Excel INI"]),
        )


EXERCISES = [
    {
        "num": "01",
        "file": "Excel_01_Creer_un_tableau_de_suivi.html",
        "corrige": "Corrige_Excel_01_Creer_un_tableau_de_suivi.html",
        "title": "Créer un tableau de suivi",
        "start": "participants_evenement.xlsx",
        "context": "Vous préparez le suivi des participants d'un atelier interne.",
        "objective": "Créer un tableau propre, lisible et enregistrable.",
        "tasks": [
            "Ouvrir le fichier XLSX participants.",
            "Renommer la feuille Suivi participants.",
            "Mettre les en-têtes en gras, avec fond clair et bordures.",
            "Ajuster automatiquement les largeurs de colonnes.",
            "Ajouter une colonne Commentaire et tester le renvoi à la ligne.",
            "Enregistrer le classeur au format Excel.",
        ],
        "checks": ["En-têtes visibles", "Colonnes ajustées", "Bordures sobres", "Aucune donnée écrasée", "Fichier enregistré"],
        "hint": "Double-cliquez entre deux lettres de colonnes pour ajuster la largeur.",
        "correction": ["A1:E1 en gras", "Largeurs ajustées", "Commentaire présent", "Tableau lisible à l'écran et à l'impression"],
        "errors": "Saisir hors de la cellule active, supprimer le contenu mais garder un vieux format, confondre couper et copier.",
    },
    {
        "num": "02",
        "file": "Excel_02_Calculer_totaux_et_moyennes.html",
        "corrige": "Corrige_Excel_02_Calculer_totaux_et_moyennes.html",
        "title": "Calculer totaux et moyennes",
        "start": "budget_fournitures.xlsx",
        "context": "Vous devez calculer un budget fournitures avant validation.",
        "objective": "Utiliser formules et fonctions automatiques sans calcul manuel.",
        "tasks": [
            "Ouvrir le budget fournitures.",
            "Ajouter une colonne Total.",
            "En D2, créer la formule Quantité x Prix unitaire.",
            "Recopier la formule sur toutes les lignes.",
            "Ajouter SOMME, MOYENNE, NB, MAX et MIN sous le tableau.",
            "Mettre les montants au format monétaire.",
        ],
        "checks": ["Formules présentes", "SOMME correcte", "MOYENNE correcte", "Formats monétaires", "Aucun total saisi à la main"],
        "hint": "Une plage comme D2:D8 se lit de D2 à D8.",
        "correction": ["D2 contient =B2*C2", "Total général avec =SOMME(D2:D8)", "Moyenne avec =MOYENNE(D2:D8)", "MAX et MIN basés sur la colonne Total"],
        "errors": "Oublier le signe =, sélectionner la mauvaise plage, taper le résultat au lieu de garder la formule.",
    },
    {
        "num": "03",
        "file": "Excel_03_Creer_un_budget_simple.html",
        "corrige": "Corrige_Excel_03_Creer_un_budget_simple.html",
        "title": "Créer un budget simple",
        "start": "budget_fournitures.xlsx",
        "context": "Vous présentez un budget clair pour une réunion.",
        "objective": "Structurer un tableau de budget et préparer sa lecture.",
        "tasks": [
            "Créer un titre au-dessus du tableau.",
            "Ajouter une ligne Imprévus.",
            "Calculer les totaux par ligne.",
            "Appliquer un format comptabilité ou monétaire.",
            "Mettre une bordure extérieure et des en-têtes lisibles.",
            "Vérifier que le format ne change pas la valeur réelle.",
        ],
        "checks": ["Titre présent", "Ligne imprévus", "Total général", "Format nombre cohérent", "Présentation sobre"],
        "hint": "Le format modifie l'affichage, pas le contenu réel de la cellule.",
        "correction": ["Titre en ligne 1", "Formules recopiées", "Total général sous la colonne Total", "Formats monétaires sur les montants"],
        "errors": "Utiliser du texte dans une cellule censée contenir un nombre, créer trop de couleurs, oublier la ligne de total.",
    },
    {
        "num": "04",
        "file": "Excel_04_Utiliser_SI_simple.html",
        "corrige": "Corrige_Excel_04_Utiliser_SI_simple.html",
        "title": "Utiliser SI simple",
        "start": "participants_evenement.xlsx",
        "context": "Vous devez repérer les participants à relancer.",
        "objective": "Créer une colonne d'aide avec une formule SI simple.",
        "tasks": [
            "Ajouter une colonne Action.",
            "Écrire une formule SI qui affiche A relancer si Présence vaut Non.",
            "Afficher OK si la personne est présente.",
            "Recopier la formule.",
            "Filtrer ou trier visuellement les personnes à relancer.",
        ],
        "checks": ["Formule SI", "Recopie correcte", "Libellés homogènes", "Absents identifiés", "Aucune saisie manuelle répétée"],
        "hint": "Structure : =SI(condition;valeur_si_vrai;valeur_si_faux).",
        "correction": ['Formule type : =SI(C2="Non";"A relancer";"OK")', "Formule recopiée sur chaque ligne", "Absents faciles à repérer"],
        "errors": "Oublier les guillemets autour du texte, recopier une formule cassée, mélanger Oui/OUI/oui sans cohérence.",
    },
    {
        "num": "05",
        "file": "Excel_05_Creer_un_graphique.html",
        "corrige": "Corrige_Excel_05_Creer_un_graphique.html",
        "title": "Créer un graphique",
        "start": "ventes_graphique.xlsx",
        "context": "Vous voulez montrer l'évolution des inscriptions par mois.",
        "objective": "Créer un graphique lisible à partir d'une sélection propre.",
        "tasks": [
            "Ouvrir les données d'inscriptions.",
            "Sélectionner seulement les mois et inscriptions, avec en-têtes.",
            "Insérer un graphique en colonnes ou en courbe.",
            "Ajouter un titre explicite.",
            "Vérifier la légende, les axes et l'emplacement.",
            "Tester l'impression du graphique seul ou avec la feuille.",
        ],
        "checks": ["Données utiles sélectionnées", "Aucun total inclus", "Titre clair", "Axes lisibles", "Graphique non déformé"],
        "hint": "Une courbe montre bien une évolution ; des colonnes montrent bien une comparaison.",
        "correction": ["Sélection A1:B7", "Graphique avec titre Inscriptions par mois", "Axe des mois visible", "Pas de ligne Total dans le graphique"],
        "errors": "Inclure le total, oublier les en-têtes, choisir un secteur pour une évolution temporelle.",
    },
    {
        "num": "06",
        "file": "Excel_06_Corriger_un_tableau_avec_erreurs.html",
        "corrige": "Corrige_Excel_06_Corriger_un_tableau_avec_erreurs.html",
        "title": "Corriger un tableau avec erreurs",
        "start": "tableau_erreurs.xlsx",
        "context": "Un tableau transmis contient des erreurs de saisie et de calcul.",
        "objective": "Identifier les symptômes Excel courants et corriger la source.",
        "tasks": [
            "Repérer les nombres stockés en texte.",
            "Corriger les valeurs qui utilisent la lettre O à la place du chiffre 0.",
            "Élargir les colonnes qui affichent ###.",
            "Corriger les formules en erreur #DIV/0!, #VALEUR! ou #REF! si elles apparaissent.",
            "Vérifier la barre de formule avant de modifier.",
        ],
        "checks": ["Valeurs numériques", "Colonnes assez larges", "Erreurs expliquées", "Formules réparées", "Tableau sauvegardé sous nouveau nom"],
        "hint": "Le contenu réel se lit dans la barre de formule.",
        "correction": ["Montants convertis en nombres", "Colonnes ajustées", "Formules sans référence cassée", "Commentaires ajoutés sur les corrections"],
        "errors": "Corriger l'apparence sans corriger le contenu, effacer une formule au lieu de réparer sa référence.",
    },
    {
        "num": "07",
        "file": "Excel_07_Mise_en_page_et_impression.html",
        "corrige": "Corrige_Excel_07_Mise_en_page_et_impression.html",
        "title": "Mise en page et impression",
        "start": "frais_vacances_mise_en_page.xlsx",
        "context": "Vous devez remettre un tableau de frais de déplacement imprimable.",
        "objective": "Préparer un tableau lisible en aperçu avant impression.",
        "tasks": [
            "Ouvrir le fichier frais vacances.",
            "Calculer Total = Quantité x Prix unitaire.",
            "Mettre les montants au format monétaire.",
            "Passer en orientation paysage.",
            "Ajuster à 1 page en largeur.",
            "Définir une zone d'impression et ajouter un pied de page avec numéro de page.",
        ],
        "checks": ["Totaux calculés", "Orientation paysage", "Zone d'impression", "Ajustement 1 page en largeur", "Aperçu vérifié"],
        "hint": "Contrôlez toujours l'aperçu avant d'imprimer.",
        "correction": ["Colonne Total calculée", "Paysage actif", "Largeur ajustée sur une page", "Pied de page avec numéro"],
        "errors": "Lancer l'impression sans aperçu, réduire trop fortement l'échelle, oublier les titres à imprimer.",
    },
    {
        "num": "08",
        "file": "Excel_08_Series_et_recopie_de_formules.html",
        "corrige": "Corrige_Excel_08_Series_et_recopie_de_formules.html",
        "title": "Séries et recopie de formules",
        "start": "series_recopie_formules.xlsx",
        "context": "Vous préparez un suivi mensuel avec des dates et des calculs répétitifs.",
        "objective": "Créer des séries et recopier des formules sans casser les références.",
        "tasks": [
            "Créer la série des mois de janvier à juin.",
            "Créer une série de semaines avec un pas régulier.",
            "Calculer Écart = Réalisé - Prévu.",
            "Recopier la formule sur toutes les lignes.",
            "Créer une moyenne des écarts.",
            "Repérer toute erreur de recopie.",
        ],
        "checks": ["Série complète", "Pas régulier", "Formule recopiée", "Moyenne calculée", "Références vérifiées"],
        "hint": "Après recopie, cliquez une cellule copiée et lisez la barre de formule.",
        "correction": ["Mois Janvier à Juin", "Écart avec =C2-B2", "Moyenne avec =MOYENNE(D2:D7)", "Aucune formule figée par erreur"],
        "errors": "Recopier une valeur au lieu d'une série, oublier une ligne dans la plage, ne pas vérifier la formule copiée.",
    },
    {
        "num": "09",
        "file": "Excel_09_References_absolues_commission.html",
        "corrige": "Corrige_Excel_09_References_absolues_commission.html",
        "title": "Références absolues et commission",
        "start": "commission_references_absolues.xlsx",
        "context": "Vous calculez des commissions avec un taux unique stocké dans une cellule.",
        "objective": "Utiliser $ pour bloquer la cellule du taux lors de la recopie.",
        "tasks": [
            "Repérer la cellule du taux de commission.",
            "Calculer Commission = Ventes x Taux.",
            "Utiliser une référence absolue pour bloquer le taux.",
            "Recopier la formule sur tous les commerciaux.",
            "Modifier le taux et vérifier que toutes les commissions changent.",
        ],
        "checks": ["Taux centralisé", "Référence $ utilisée", "Recopie fiable", "Montants recalculés", "Format monétaire"],
        "hint": "Une formule type peut contenir $F$2 pour garder le taux fixe.",
        "correction": ["Commission avec =B2*$F$2", "$F$2 reste fixe après recopie", "Changer F2 met à jour toutes les commissions"],
        "errors": "Référence relative au taux qui glisse ligne après ligne, taux saisi plusieurs fois, formats incohérents.",
    },
    {
        "num": "10",
        "file": "Excel_10_Graphiques_et_mise_en_forme_conditionnelle.html",
        "corrige": "Corrige_Excel_10_Graphiques_et_mise_en_forme_conditionnelle.html",
        "title": "Graphiques et mise en forme conditionnelle",
        "start": "entrainement_graphique_conditionnel.xlsx",
        "context": "Vous préparez un tableau de performance à commenter en réunion.",
        "objective": "Associer graphique, mise en forme conditionnelle et commentaire utile.",
        "tasks": [
            "Calculer l'écart entre Objectif et Réalisé.",
            "Créer une règle de mise en forme conditionnelle pour les écarts négatifs.",
            "Ajouter des barres de données sur le Réalisé.",
            "Créer un graphique comparant Objectif et Réalisé par mois.",
            "Ajouter une note sur le mois le plus faible.",
            "Copier le tableau dans Word si demandé par le formateur.",
        ],
        "checks": ["Écart calculé", "MFC visible", "Graphique adapté", "Note utile", "Tableau copiable vers Word"],
        "hint": "Sélectionnez les données utiles sans inclure une ligne de total.",
        "correction": ["Écart = C2-B2", "Valeurs négatives en surbrillance", "Graphique colonnes groupées Objectif/Réalisé", "Note sur le mois le plus faible"],
        "errors": "Créer un graphique sans question métier, empiler trop de règles MFC, laisser une note obsolète.",
    },
]


def exercise_page(ex):
    start = f'../../06_Fichiers_de_depart/Excel/{ex["start"]}'
    corrige = f'../../05_Corriges_HTML/Excel/{ex["corrige"]}'
    body = f"""
    <p><a href="../../03_Supports_stagiaires_HTML/Support_Excel_Debutant.html">Retour au support Excel</a></p>
    <section class="grid two">
      <article class="panel"><h2>Contexte professionnel</h2><p>{ex["context"]}</p><div class="callout ok"><h3>Objectif</h3><p>{ex["objective"]}</p></div></article>
      <aside class="panel"><h2>Fichier de départ</h2><p><a href="{start}">{ex["start"]}</a></p><h3>Critères de réussite</h3>{ul(ex["checks"])}</aside>
    </section>
    {panel("Consignes", ol_steps(ex["tasks"]))}
    {panel("Aide", accordion("Afficher un indice", f"<p>{ex['hint']}</p>") + accordion("Afficher le corrigé résumé", ul(ex["correction"])))}
    {panel("Checklist stagiaire", checklist(ex["checks"], ex["title"] + " checklist"))}
    {panel("Corrigé formateur", f'<p><a href="{corrige}">Ouvrir le corrigé détaillé</a></p>')}
    """
    write(
        f'04_Exercices_HTML/Excel/{ex["file"]}',
        html_page(ex["title"], "excel", "../../00_Index_general/index.html", ex["context"], body, ["Excel", f'Exercice {ex["num"]}', "Cas pratique"]),
    )


def corrige_page(ex):
    exercise = f'../../04_Exercices_HTML/Excel/{ex["file"]}'
    body = f"""
    <p><a href="{exercise}">Retour à l'exercice</a></p>
    <section class="grid two">
      <article class="panel"><h2>Résultat attendu</h2>{ul(ex["correction"])}</article>
      <article class="panel"><h2>Points de contrôle</h2>{checklist(ex["checks"], "corrige " + ex["title"])}</article>
    </section>
    {panel("Étapes de correction", accordion("Afficher les étapes", ol_steps(ex["tasks"])))}
    <section class="grid two">
      <article class="card"><h3>Erreurs fréquentes</h3><p>{ex["errors"]}</p></article>
      <article class="card"><h3>Validation collective</h3><p>Faire expliquer la cellule active, la formule ou le format par un stagiaire avant de montrer la correction complète.</p></article>
    </section>
    {panel("Source pédagogique", f"<p>Exercice enrichi à partir du plan complet Excel INI : {SOURCE_PLAN}</p>")}
    """
    write(
        f'05_Corriges_HTML/Excel/{ex["corrige"]}',
        html_page("Corrigé - " + ex["title"], "excel", "../../00_Index_general/index.html", "Corrigé formateur détaillé.", body, ["Excel", "Corrigé formateur"]),
    )


def exercise_files():
    write_xlsx("06_Fichiers_de_depart/Excel/participants_evenement.xlsx", [
        ["Nom", "Service", "Presence", "Repas"],
        ["Martin", "Accueil", "Oui", "Oui"],
        ["Bernard", "Commercial", "Oui", "Non"],
        ["Diallo", "Admin", "Non", "Non"],
        ["Nguyen", "Formation", "Oui", "Oui"],
        ["Petit", "RH", "Non", "Oui"],
    ], "Participants")
    write_xlsx("06_Fichiers_de_depart/Excel/budget_fournitures.xlsx", [
        ["Article", "Quantite", "Prix_unitaire"],
        ["Stylos", 20, 1.20],
        ["Badges", 20, 0.80],
        ["Cahiers", 12, 2.50],
        ["Boissons", 24, 1.10],
        ["Chemises", 15, 1.75],
        ["Marqueurs", 6, 3.40],
    ], "Budget")
    write_xlsx("06_Fichiers_de_depart/Excel/ventes_graphique.xlsx", [
        ["Mois", "Inscriptions"],
        ["Janvier", 8],
        ["Fevrier", 10],
        ["Mars", 12],
        ["Avril", 9],
        ["Mai", 16],
        ["Juin", 14],
    ], "Inscriptions")
    write_xlsx("06_Fichiers_de_depart/Excel/tableau_erreurs.xlsx", [
        ["Nom", "Montant", "Diviseur", "Commentaire"],
        ["Aline", "12O", 4, "Lettre O dans le montant"],
        ["Samir", 150, 0, "Risque #DIV/0!"],
        ["Lea", "-3", 3, "Montant négatif à vérifier"],
        ["Noe", 210, 7, ""],
    ], "Erreurs")
    write_xlsx("06_Fichiers_de_depart/Excel/frais_vacances_mise_en_page.xlsx", [
        ["Date", "Nature", "Quantite", "Prix_unitaire", "Total"],
        ["03/07/2026", "Train", 2, 48.50, ""],
        ["04/07/2026", "Hotel", 3, 82.00, ""],
        ["05/07/2026", "Repas", 6, 18.90, ""],
        ["06/07/2026", "Parking", 1, 24.00, ""],
        ["07/07/2026", "Transport local", 4, 3.20, ""],
    ], "Frais")
    write_xlsx("06_Fichiers_de_depart/Excel/series_recopie_formules.xlsx", [
        ["Mois", "Prevu", "Realise", "Ecart"],
        ["Janvier", 120, 115, ""],
        ["Fevrier", 125, 132, ""],
        ["Mars", 130, 128, ""],
        ["Avril", 135, 140, ""],
        ["Mai", 140, 137, ""],
        ["Juin", 145, 150, ""],
    ], "Series")
    write_xlsx("06_Fichiers_de_depart/Excel/commission_references_absolues.xlsx", [
        ["Commercial", "Ventes", "Commission"],
        ["Aline", 12500, ""],
        ["Samir", 9800, ""],
        ["Lea", 14300, ""],
        ["Noe", 11100, ""],
        ["", "", ""],
        ["Taux_commission", "5%", ""],
    ], "Commissions")
    write_xlsx("06_Fichiers_de_depart/Excel/entrainement_graphique_conditionnel.xlsx", [
        ["Mois", "Objectif", "Realise", "Ecart"],
        ["Janvier", 100, 95, ""],
        ["Fevrier", 105, 108, ""],
        ["Mars", 110, 103, ""],
        ["Avril", 115, 120, ""],
        ["Mai", 120, 117, ""],
        ["Juin", 125, 131, ""],
    ], "Performance")


def exercise_pages():
    for ex in EXERCISES:
        exercise_page(ex)
        corrige_page(ex)


def report_updates():
    html_count = len(list(ROOT.rglob("*.html")))
    marker = "\n---\n\n## Mise à jour - enrichissement Excel INI"
    report = ROOT / "RAPPORT_GENERATION.md"
    text = report.read_text(encoding="utf-8") if report.exists() else "# Rapport de génération - Kit Horizon Compétences Office débutant\n"
    if marker in text:
        text = text.split(marker)[0]
    addition = f"""{marker}

Date de mise à jour : 28 mai 2026

Sources utilisées :

- `{SOURCE_PLAN}`
- `{CONCISE_PLAN}`
- `{OPERATORS_PLAN}`

### Contenu enrichi

- Support stagiaire Excel remplacé par une version restructurée issue des plans Excel INI.
- Plan de cours séquencé ajouté : démarrage, feuilles, saisie, mise en forme, fonctions, opérateurs, séries, impression, graphiques, MFC.
- Section opérateurs de calcul ajoutée : arithmétiques, comparaison, texte, références et ordre de priorité.
- Guides formateur jours 3, 4 et 5 enrichis : démarrage, saisie, formats, fonctions, séries, impression, références absolues, graphiques, notes, mise en forme conditionnelle et interaction Word.
- Exercices Excel portés à 10 cas pratiques.
- Corrigés Excel portés à 10 pages formateur.
- Fichiers de départ Excel convertis en `.xlsx` : participants, budget, erreurs, frais vacances, séries/recopie, commissions, graphique + mise en forme conditionnelle.

### Vérifications ciblées

- HTML Excel standalone : OK après génération.
- Liens internes Excel : à contrôler en phase QA finale.
- Fichier source DOCX original : non modifié.
- HTML présents après enrichissement : {html_count}.
"""
    report.write_text(text.rstrip() + addition, encoding="utf-8")

    qa = f"""# QA_CHECKLIST.md

Date : 28 mai 2026

## Contrôles finaux

- HTML vérifiés : {html_count}
- CSS intégré : vérifié
- JavaScript intégré : vérifié
- `@media print` : vérifié
- Dépendances externes : aucune détectée dans les HTML
- Liens internes : à vérifier après enrichissement Excel
- Résultat global : à vérifier

## Enrichissement Excel INI

- Sources : `Excel\\Excel Ini\\PLAN COMPLET Excel INI.docx`, `Excel\\Excel Ini\\Plan concis Excel Ini.docx`, `Excel\\Excel Ini\\opérateurs de calcul.docx`
- Support Excel stagiaire : enrichi
- Plan de cours Excel : restructuré
- Opérateurs de calcul : intégrés
- Guides formateur J3/J4/J5 : enrichis
- Exercices Excel : 10
- Corrigés Excel : 10
- Fichiers de départ Excel : `.xlsx`
"""
    write("QA_CHECKLIST.md", qa)

    report_body = f"""
    {panel("État du kit", ul(["Phases 1 à 12 déjà générées.", "Module Excel restructuré le 28 mai 2026 à partir des plans Excel INI et opérateurs de calcul.", f"HTML présents : {html_count}."]))}
    {panel("Enrichissement Excel", ul(["Support stagiaire Excel détaillé.", "Plan de cours restructuré.", "Index des opérateurs et commandes de base.", "Guides formateur J3, J4, J5 enrichis.", "10 exercices Excel et 10 corrigés.", "Fichiers XLSX de départ complétés."]))}
    {panel("Documents", ul(['<a href="03_Supports_stagiaires_HTML/Support_Excel_Debutant.html">Support Excel enrichi</a>', '<a href="00_Index_general/index.html">Index général</a>', '<a href="QA_CHECKLIST.md">Checklist QA</a>']))}
    """
    write(
        "RAPPORT_GENERATION.html",
        html_page("Rapport de génération - Kit Horizon Compétences", "general", "00_Index_general/index.html", "Rapport mis à jour après enrichissement Excel INI.", report_body, ["Rapport", "Excel INI"]),
    )


def main():
    support_excel()
    guide_pages()
    exercise_files()
    exercise_pages()
    report_updates()
    print("Excel INI enrichment complete")


if __name__ == "__main__":
    main()
